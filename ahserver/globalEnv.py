# -*- coding:utf8 -*-
import os
import sys
import codecs 
from urllib.parse import quote
import json

import random
import time
import datetime
from openpyxl import Workbook
from tempfile import mktemp

from appPublic.jsonConfig import getConfig
from appPublic.Singleton import GlobalEnv
from appPublic.argsConvert import ArgsConvert
from appPublic.timeUtils import str2Date,str2Datetime,curDatetime,getCurrentTimeStamp,curDateString, curTimeString
from appPublic.folderUtils import folderInfo
from appPublic.uniqueID import setNode,getID
from appPublic.unicoding import unicoding,uDict,uObject
from appPublic.Singleton import SingletonDecorator

from sqlor.dbpools import DBPools,runSQL,runSQLPaging
from sqlor.crud import CRUD


from .xlsxData import XLSXData
from .uriop import URIOp

from .serverenv import ServerEnv

def data2xlsx(rows,headers=None):
	wb = Workbook()
	ws = wb.active

	i = 1
	if headers is not None:
		for j in range(len(headers)):
			v = headers[j].title if headers[j].get('title',False) else headers[j].name
			ws.cell(column=j+1,row=i,value=v)
		i += 1
	for r in rows:
		for j in range(len(r)):
			v = r[headers[j].name]
			ws.cell(column=j+1,row=i,value=v)
		i += 1
	name = mktemp(suffix='.xlsx')
	wb.save(filename = name)
	wb.close()
	return name
	
class FileOutZone(Exception):
	def __init__(self,fp,*args,**kwargs):
		super(FileOutZone,self).__init__(*args,**kwargs)
		self.openfilename = fp
	
	def __str__(self):
		return self.openfilename + ': not allowed to open'
		
def openfile(url,m):
	fp = abspath(url)
	if fp is None:
		print('openfile(',url,m,'),url is not match a file')
		raise Exception('url can not mathc a file')
	config = getConfig()
	paths = [ os.path.abspath(p) for p in config.website.paths ]
	fs = config.get('allow_folders',[])
	fs = [ os.path.abspath(i) for i in fs + paths ]
	r = False
	for f in fs:
		if fp.startswith(f):
			r = True
			break
	if not r:
		raise FileOutZone(fp)
	return open(fp,m)
	
def isNone(a):
	return a is None

def abspath(path):
	config = getConfig()
	paths = [ os.path.abspath(p) for p in config.website.paths ]
	for root in paths:
		p = root + path
		if os.path.exists(root+path):
			return p
			
	return None

def appname():
	config = getConfig()
	try:
		return config.license.app
	except:
		return "test app"
	
def configValue(ks):
	config = getConfig()
	try:
		a = eval('config' + ks)
		return a
	except:
		return None

def visualcoding():
	return configValue('.website.visualcoding');

def file_download(request,path,name,coding='utf8'):
	f = openfile(path,'rb')
	b = f.read()
	f.close()
	fname = quote(name).encode(coding)
	hah = b"attachment; filename=" + fname
	# print('file head=',hah.decode(coding))
	request.setHeader(b'Content-Disposition',hah)
	request.setHeader(b'Expires',0)
	request.setHeader(b'Cache-Control',b'must-revalidate, post-check=0, pre-check=0')
	request.setHeader(b'Content-Transfer-Encoding',b'binary')
	request.setHeader(b'Pragma',b'public')
	request.setHeader(b'Content-Length',len(b))
	request.write(b)
	request.finish()
	
def initEnv():
	pool = DBPools()
	g = ServerEnv()
	g.configValue = configValue
	g.visualcoding = visualcoding
	g.uriop = URIOp
	g.isNone = isNone
	g.json = json
	g.int = int
	g.str = str
	g.float = float
	g.type = type
	g.ArgsConvert = ArgsConvert
	g.time = time
	g.curDateString = curDateString
	g.curTimeString = curTimeString
	g.datetime = datetime
	g.random = random
	g.str2date = str2Date
	g.str2datetime = str2Datetime
	g.curDatetime = curDatetime
	g.uObject = uObject
	g.uuid = getID
	g.runSQL = runSQL
	g.runSQLPaging = runSQLPaging
	g.runSQLIterator = pool.runSQL
	g.runSQLResultFields = pool.runSQLResultFields
	g.getTables = pool.getTables
	g.getTableFields = pool.getTableFields
	g.getTablePrimaryKey = pool.getTablePrimaryKey
	g.getTableForignKeys = pool.getTableForignKeys
	g.folderInfo = folderInfo
	g.abspath = abspath
	g.data2xlsx = data2xlsx
	g.xlsxdata = XLSXData
	g.openfile = openfile
	g.CRUD = CRUD
	g.DBPools = DBPools
